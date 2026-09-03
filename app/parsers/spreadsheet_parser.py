"""CSV/XLSX parser that preserves sheet, column and row provenance."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Generator

from openpyxl import load_workbook

from app.parsers.base import BaseParser, ParserPrelude, StructuredBlock


def _cell_text(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def _rows_to_blocks(sheet_name: str, rows: list[list[str]], *, group_size: int = 30) -> list[StructuredBlock]:
    meaningful_rows = [row for row in rows if any(cell for cell in row)]
    if not meaningful_rows:
        return []
    headers = meaningful_rows[0]
    header_text = " | ".join(headers)
    blocks = [
        StructuredBlock(
            text=f"工作表: {sheet_name}\n列: {header_text}",
            kind="table_schema",
            section_title=sheet_name,
            heading_path=(sheet_name,),
            table_headers=tuple(headers),
            metadata={"sheet_name": sheet_name, "row_start": 1, "row_end": 1},
        )
    ]
    data_rows = meaningful_rows[1:]
    for offset in range(0, len(data_rows), group_size):
        group = data_rows[offset: offset + group_size]
        row_start = offset + 2
        row_end = row_start + len(group) - 1
        text = "\n".join([header_text, *(" | ".join(row) for row in group)])
        blocks.append(
            StructuredBlock(
                text=text,
                kind="table",
                section_title=sheet_name,
                heading_path=(sheet_name,),
                table_headers=tuple(headers),
                metadata={"sheet_name": sheet_name, "row_start": row_start, "row_end": row_end},
            )
        )
    return blocks


class SpreadsheetParser(BaseParser):
    def get_prelude(self, file_path: Path) -> ParserPrelude:
        return ParserPrelude(title=file_path.stem)

    def parse_blocks(self, file_path: Path) -> list[StructuredBlock]:
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            with open(file_path, encoding="utf-8-sig", errors="replace", newline="") as handle:
                rows = [[_cell_text(cell) for cell in row] for row in csv.reader(handle)]
            return _rows_to_blocks(file_path.stem, rows)

        workbook = load_workbook(file_path, read_only=True, data_only=False)
        try:
            blocks: list[StructuredBlock] = []
            for worksheet in workbook.worksheets:
                rows = [
                    [_cell_text(value) for value in row]
                    for row in worksheet.iter_rows(values_only=True)
                ]
                blocks.extend(_rows_to_blocks(worksheet.title, rows))
            return blocks
        finally:
            workbook.close()

    def parse(self, file_path: Path) -> str:
        return "\n\n".join(block.text for block in self.parse_blocks(file_path))

    def parse_stream(self, file_path: Path) -> Generator[str, None, None]:
        for block in self.parse_blocks(file_path):
            yield block.text
